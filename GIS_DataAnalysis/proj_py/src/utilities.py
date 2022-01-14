'''
========================
utilities.py
========================
Created on Oct.9, 2019
@author: Xu Ronghua
@Email:  rxu22@binghamton.edu
@TaskDescription: This module provide utility function to support project.
@Reference: 
'''

import glob, os, sys
import pandas as pd
from openpyxl import load_workbook
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter, MaxNLocator
import numpy as np
import csv

csv.field_size_limit(sys.maxsize)

'''
FileUtil class for handling file read and write
'''
class FileUtil(object):
	@staticmethod
	def list_files(dir_path="./", reg_str='*'):
		'''
		Function: list all files in dir_path directory based on reg_str condition
		@arguments: 
		(in)    dir_path:   directory path for file search. set current path as default directory
		(in)    reg_str:   	input regex string to filter files, eg, *.log. list all files as default
		(out)   ls_files:   	return listed files
		'''
		# save pwd for restore path
		pwd = os.getcwd()
		os.chdir(dir_path)
		ls_files=[]
		for file in glob.glob(reg_str):
			ls_files.append(file)
		
		# restore original path
		os.chdir(pwd)
		return ls_files

	@staticmethod
	def AddDataByLine(filepath, ls_data):
		'''
		Function: write list data to file by each line
		@arguments: 
		(in)  filepath:   	input file path
		(in)  ls_data:   	list data for writing
		'''
		#define file handle to open select file for write data
		fname = open(filepath, 'w') 
		
		#for each lines in data and write to file
		for linedata in ls_data:
			#write line data to file
			fname.write("%s\n" %(linedata))
		
		#close file
		fname.close()

	@staticmethod
	def xlsread(filepath, dataRange=[], is_dropNaN=False):
		'''
		Function: Read data from excel
		@arguments: 
		(in)  filepath:   		input file path
		(in)  dataRange:   		data range
		(in)  is_dropNaN:   	remove rows containing NaN
		(out) ls_dataset:   	return line list object
		'''

		if( dataRange==[] ):
			df = pd.read_excel(filepath, sheet_name='Sheet1')

		else:
			wb = load_workbook(filepath, read_only=True)
			ws = wb['Sheet1']
			data_rows = []
			# Read the cell values into a list of lists
			for row in ws[dataRange[0]:dataRange[1]]:
				data_cols = []
				for cell in row:
					data_cols.append(cell.value)
				data_rows.append(data_cols)
			# Transform into dataframe
			df = pd.DataFrame(data_rows)
		# if drop nan rows
		if(is_dropNaN):
			df = df.dropna()

		# Read the cell values into a list of lists
		ls_dataset=[]
		rows_id = 0
		for rows_id in range(0, df.shape[0]):
			cols_id=0
			data_cols = []
			for cols_id in range(0, df.shape[1]):
				data_cols.append(df.iloc[rows_id,cols_id])
			ls_dataset.append(data_cols)
		'''print(df.columns)
		print(df[df.columns[0:2]])
		print(dataset[0:10])'''
		return ls_dataset

	@staticmethod
	def csv_read(csv_file):
		'''
		Function: Read data from csv file
		@arguments: 
		(out) np_dataset:   	return np.array object
		(in) csv_file:   		csv file path
		'''
		ls_dataset = []
		with open(csv_file, 'r') as csvFile:
			csv_reader = csv.reader(csvFile, delimiter=',')
			for row in csv_reader:
				ls_dataset.append(row)
		# transfer to np array and return
		np_dataset = np.array(ls_dataset)
		return np_dataset

	@staticmethod
	def csv_write(csv_file, dataset):
		'''
		Function: Write data to csv file
		@arguments: 
		(in) dataset:   	np.array data
		(in) csv_file:   	csv file path
		'''
		ls_dataset = []
		with open(csv_file, 'w') as csvFile:
			csv_writer = csv.writer(csvFile, delimiter=',')
			for row in dataset:
				csv_writer.writerow(row)

	@staticmethod
	def data2csv(csv_Path, np_dataset):
		'''
		Function: Write data to csv file
		@arguments: 
		(in) np_dataset:   	list matrix
		(in) csv_Path:   	csv file path
		'''
		ls_dataset = TypesUtil.np2list(np_dataset)
		pdData = pd.DataFrame(ls_dataset)
		pdData.to_csv(csv_Path, index=None, header=None)

	@staticmethod
	def csv2data(csv_Path):
		'''
		Function: Read data from csv file
		@arguments: 
		(in) csv_Path:   	csv file path
		(out) np_dataset:	return np data
		'''
		raw_dataset = []
		with open(csv_Path, 'r') as csvFile:
			csv_reader = csv.reader(csvFile, delimiter=',')
			for row in csv_reader:
				raw_dataset.append(row)
	
		## for each dimension to rebuild array.
		ls_dataset = []
		for d1_dataset in raw_dataset:
			ls_d1 = []
			for d2_dataset in d1_dataset:
				ls_d2 = []
				for row_data in d2_dataset.split("],"):
					ls_row = row_data.replace("[", "").replace("]", "").replace(" ", "").split(",")
					## append row array list to ls_d2
					ls_d2.append(ls_row)
				## append (row,col) array list to ls_d1 (data sheet)
				ls_d1.append(ls_d2)
			## append ls_d1 (data sheet) to ls_dataset.
			ls_dataset.append(ls_d1)
		## transfer ls_dataset to np array and return
		np_dataset = TypesUtil.list2np(ls_dataset)
		return np_dataset

'''
TypesUtil class for data type format transfer
'''
class TypesUtil(object):
	# list dataset to numpy matrix
	@staticmethod
	def list2np(ls_data):
		# transfer to np array and return
		np_data = np.array(ls_data, dtype=np.float32)
		return np_data

	# numpy matrix to list dataset
	@staticmethod
	def np2list(np_data):
		# transfer to list dataset and return
		ls_data = np_data.tolist()
		return ls_data

'''
DataUtil class for handling data preparation
'''
class DataUtil(object):
	@staticmethod
	def PrepareData(ls_dataset):
		'''
		Function: pre-process data for furthre calculation
		@arguments: 
		(in) ls_dataset:   		return line list object
		(out) pre_dataset:   	return line list object
		'''
		pre_dataset = []

		#calculate year median
		year_sum=1;
		value_sum = ls_dataset[0][0];
		preYear = ls_dataset[0][1];
		year_mdeian=0.0;
		#print(value_sum, ',', preYear)
		i=0
		for i in range(1,len(ls_dataset)):
			tmp_data = []
			if( preYear == ls_dataset[i][1]):
				year_sum = year_sum + 1
				value_sum = value_sum + ls_dataset[i][0]
			else:
				year_mdeian=value_sum/year_sum;
				#print([year_mdeian, value_sum, year_sum]);
				tmp_data.append(preYear)
				tmp_data.append(year_mdeian)
				pre_dataset.append(tmp_data)
				#initialize year_sum, value_sum and update preYear
				year_sum = 1;
				value_sum = ls_dataset[i][0];
				preYear = ls_dataset[i][1];			

		return pre_dataset

'''
PlotUtil class for data visualization
'''
class PlotUtil(object):
	@staticmethod
	def PlotData(ls_dataset, x_label, y_label, 
				font_size=14, is_show=True, is_savefig=False, datafile=''):
		'''
		Function: plot data on fig
		@arguments: 
		(in) ls_dataset:   		return line list object
		'''
		#generate x and y data
		xdata = [];
		ydata = [];

		for i in range(0, len(ls_dataset)):
			xdata.append(ls_dataset[i][0])
			ydata.append(ls_dataset[i][1])

		matplotlib.rcParams.update({'font.size': 12})
		fig, ax = plt.subplots()
		ax.xaxis.set_major_locator(MaxNLocator(integer=True))
		ax.plot(xdata, ydata, '.')
		#ax.set_xticklabels(xdata, fontsize=14)
		plt.xlabel(x_label, fontsize=font_size)
		plt.ylabel(y_label, fontsize=font_size) 
		plt.ylim(-0.05, 1.05)
		
		if( is_show ):
			plt.show()
		if( is_savefig ):
			## set ext for figure
			figname = os.path.splitext(datafile)[0] +'.png'

			## check if parent directory is existed.
			png_dir = os.path.dirname(figname)
			if( (png_dir!='') and (not os.path.exists(png_dir)) ):
				os.makedirs(png_dir)
			
			## save figure to disk.
			fig.savefig(figname)
		plt.close()

	@staticmethod
	def Plotfit(xdata, ydata, x, y,
				 x_label, y_label, plt_title='', font_size=14, 
				 is_show=True, is_savefig=False, datafile=''):
		'''
		Function: plot data and fit curve on fig
		@arguments: 
		(in) [xdata, ydata]:   		original data
		(in) [x, y]:   				fit data: y=fit_fun(x)
		'''

		matplotlib.rcParams.update({'font.size': 12})
		fig, ax = plt.subplots()
		ax.xaxis.set_major_locator(MaxNLocator(integer=True))
		ax.plot(xdata, ydata, '.', color='b', label='data')
		#plt.scatter(xdata, ydata, s=80, facecolors='none', edgecolors='b')
		#ax.plot(xdata_opt, ydata_opt, '*', color='g', label='data')
		ax.plot(x, y, color='r', label='fit')
		plt.xlabel(x_label, fontsize=font_size)
		plt.ylabel(y_label, fontsize=font_size) 
		plt.ylim(-0.05, 1.05)
		plt.title(plt_title)
		
		if( is_show ):
			plt.show()
		if( is_savefig ):
			## set ext for figure
			figname = os.path.splitext(datafile)[0] +'.png'

			## check if parent directory is existed.
			png_dir = os.path.dirname(figname)
			if( (png_dir!='') and (not os.path.exists(png_dir)) ):
				os.makedirs(png_dir)
			
			## save figure to disk.
			fig.savefig(figname)
		plt.close()

	@staticmethod
	def Plotfit_Opt(xdata, ydata, xdata_opt, ydata_opt, x, y,
				 x_label, y_label, plt_title='', font_size=14, 
				 is_show=True, is_savefig=False, datafile=''):
		'''
		Function: plot data and fit curve on fig with optimization
		@arguments: 
		(in) [xdata, ydata]:   			original data
		(in) [xdata_opt, ydata_opt]:   	optimized data
		(in) [x, y]:   					fit data: y=fit_fun(x)
		'''

		matplotlib.rcParams.update({'font.size': 12})
		fig, ax = plt.subplots()
		ax.xaxis.set_major_locator(MaxNLocator(integer=True))
		#ax.plot(xdata, ydata, 'o', color='b', label='data')
		plt.scatter(xdata, ydata, s=80, facecolors='none', edgecolors='b')
		ax.plot(xdata_opt, ydata_opt, '*', color='g', label='data')
		ax.plot(x, y, color='r', label='fit')
		plt.xlabel(x_label, fontsize=font_size)
		plt.ylabel(y_label, fontsize=font_size) 
		plt.ylim(-0.05, 1.05)
		plt.title(plt_title)
		
		if( is_show ):
			plt.show()
		if( is_savefig ):
			## set ext for figure
			figname = os.path.splitext(datafile)[0] +'.png'

			## check if parent directory is existed.
			png_dir = os.path.dirname(figname)
			if( (png_dir!='') and (not os.path.exists(png_dir)) ):
				os.makedirs(png_dir)
			
			## save figure to disk.
			fig.savefig(figname)
		plt.close()

